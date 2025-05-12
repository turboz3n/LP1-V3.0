import os
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook

class DocumentReaderSkill:
    def describe(self):
        return {
            "name": "document_reader",
            "trigger": ["read document", "open file", "extract content"],
            "description": "Reads basic content from PDF, DOCX, or XLSX documents."
        }

    async def handle(self, user_input: str, context: dict) -> str:
        try:
            path = user_input.split()[-1]
            if not os.path.exists(path):
                return f"File not found: {path}"

            if path.endswith(".pdf"):
                text = "\n".join([p.extract_text() or "" for p in PdfReader(path).pages])
                return text[:1500] or "[No extractable text found]"

            elif path.endswith(".docx"):
                doc = Document(path)
                text = "\n".join([p.text for p in doc.paragraphs])
                return text[:1500] or "[No readable content found]"

            elif path.endswith(".xlsx"):
                wb = load_workbook(path)
                sheet = wb.active
                rows = ["\t".join([str(cell.value) for cell in row]) for row in sheet.iter_rows(min_row=1, max_row=20)]
                return "\n".join(rows)

            else:
                return "Unsupported file format. Use PDF, DOCX, or XLSX."

        except Exception as e:
            return f"[Document Reader Error] {e}"
